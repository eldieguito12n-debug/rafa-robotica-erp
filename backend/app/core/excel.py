from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="1e3a5f", size=16)
SUBTITLE_FONT = Font(name="Calibri", bold=True, color="333333", size=12)
CELL_FONT = Font(name="Calibri", size=10)
CELL_MONEY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _style_header_row(ws, row_idx: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _style_data_cells(ws, start_row: int, end_row: int, num_cols: int, money_cols=None):
    money_cols = money_cols or []
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if col in money_cols:
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '"$"#,##0.00'
            elif col == 1:
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT


def _autosize_cols(ws, num_cols: int, min_width=10, max_width=50):
    for col_idx in range(1, num_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = max_len


def generate_inventory_excel_bytes(items_list: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "ROBOLAB S.A.S. - REPORTE DE INVENTARIO"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Items: {len(items_list)}"
    sub.font = SUBTITLE_FONT
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = [
        "ID", "Nombre", "Categoría", "SKU/Código", "Cantidad",
        "Stock Mínimo", "Costo Unitario", "Valor Total", "Proveedor",
        "Ubicación", "Alerta Stock Bajo"
    ]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _style_header_row(ws, header_row, len(headers))

    start_data = header_row + 1
    current_row = start_data
    for it in items_list:
        qty = int(it.get("quantity", 0) or 0)
        min_stock = int(it.get("min_stock", 0) or 0)
        unit_cost = float(it.get("unit_cost", 0) or 0)
        category = it.get("category")
        if hasattr(category, "value"):
            category = category.value
        ws.cell(row=current_row, column=1, value=it.get("id"))
        ws.cell(row=current_row, column=2, value=it.get("name"))
        ws.cell(row=current_row, column=3, value=str(category or ""))
        ws.cell(row=current_row, column=4, value=it.get("sku") or "")
        ws.cell(row=current_row, column=5, value=qty)
        ws.cell(row=current_row, column=6, value=min_stock)
        ws.cell(row=current_row, column=7, value=unit_cost)
        ws.cell(row=current_row, column=8, value=qty * unit_cost)
        ws.cell(row=current_row, column=9, value=it.get("supplier") or "")
        ws.cell(row=current_row, column=10, value=it.get("location") or "")
        ws.cell(row=current_row, column=11, value="SÍ" if qty <= min_stock else "NO")
        current_row += 1

    _style_data_cells(ws, start_data, current_row - 1, len(headers), money_cols=[7, 8])
    _autosize_cols(ws, len(headers))

    ws.freeze_panes = "A5"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_report_excel_bytes(report_type: str, records_list: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    report_type_str = str(report_type).upper()
    ws.title = f"Reporte_{report_type_str[:15]}"

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"ROBOLAB S.A.S. - REPORTE {report_type_str}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Registros: {len(records_list)}"
    sub.font = SUBTITLE_FONT
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4

    if report_type == "general":
        headers = [
            "ID", "Fecha", "Tipo", "Descripción", "Categoría",
            "Referencia", "Método Pago", "Monto"
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=h)
        _style_header_row(ws, header_row, len(headers))

        start_data = header_row + 1
        current_row = start_data
        total_ingresos = 0.0
        total_egresos = 0.0
        for r in records_list:
            rtype = r.get("type")
            if hasattr(rtype, "value"):
                rtype = rtype.value
            amount = float(r.get("amount", 0) or 0)
            if str(rtype).lower() in ("ingreso", "venta"):
                total_ingresos += amount
            elif str(rtype).lower() in ("egreso", "compra"):
                total_egresos += amount
            ws.cell(row=current_row, column=1, value=r.get("id"))
            ws.cell(row=current_row, column=2, value=str(r.get("date", "")))
            ws.cell(row=current_row, column=3, value=str(rtype or "").upper())
            ws.cell(row=current_row, column=4, value=str(r.get("description", "")))
            ws.cell(row=current_row, column=5, value=str(r.get("category", "") or ""))
            ws.cell(row=current_row, column=6, value=str(r.get("reference", "") or ""))
            ws.cell(row=current_row, column=7, value=str(r.get("payment_method", "") or ""))
            ws.cell(row=current_row, column=8, value=amount)
            current_row += 1

        _style_data_cells(ws, start_data, current_row - 1, len(headers), money_cols=[8])

        resumen_row = current_row + 2
        ws.cell(row=resumen_row, column=7, value="TOTAL INGRESOS:").font = Font(bold=True, color="1e3a5f")
        ws.cell(row=resumen_row, column=8, value=total_ingresos).font = Font(bold=True, color="1e3a5f")
        ws.cell(row=resumen_row, column=8).number_format = '"$"#,##0.00'
        ws.cell(row=resumen_row + 1, column=7, value="TOTAL EGRESOS:").font = Font(bold=True, color="c0392b")
        ws.cell(row=resumen_row + 1, column=8, value=total_egresos).font = Font(bold=True, color="c0392b")
        ws.cell(row=resumen_row + 1, column=8).number_format = '"$"#,##0.00'
        ws.cell(row=resumen_row + 2, column=7, value="UTILIDAD:").font = Font(bold=True, color="27ae60")
        util = ws.cell(row=resumen_row + 2, column=8, value=total_ingresos - total_egresos)
        util.font = Font(bold=True, color="27ae60")
        util.number_format = '"$"#,##0.00'

    else:
        if records_list:
            headers = list(records_list[0].keys())
        else:
            headers = ["Datos"]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=str(h).replace("_", " ").title())
        _style_header_row(ws, header_row, len(headers))
        start_data = header_row + 1
        current_row = start_data
        for r in records_list:
            for col_idx, key in enumerate(headers, start=1):
                v = r.get(key, "")
                if hasattr(v, "value"):
                    v = v.value
                ws.cell(row=current_row, column=col_idx, value=v if not isinstance(v, (dict, list)) else str(v))
            current_row += 1
        _style_data_cells(ws, start_data, current_row - 1, len(headers))

    num_cols = max(ws.max_column, 1)
    _autosize_cols(ws, num_cols)
    ws.freeze_panes = f"A{header_row + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
